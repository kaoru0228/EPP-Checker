import pandas as pd
from io import BytesIO
from openpyxl.styles import PatternFill


# 返り値は，DataFrameが子の数分格納されたリスト．
# 列はマーカー，行は親候補にを表し，親候補としてありうるかの真偽値(0, 1)が入る．
def check_father(data_child, data_mother, data_father):
    markers = [col for col in data_child.columns[1:] if not 'Unnamed' in col]
    results = []

    # 子供の数分ループ
    for i in range(data_child.shape[0]):
        result = {}

        # 各マーカーについて探索
        for marker in markers:
            result[marker] = []
            marker_col = data_child.columns.get_loc(marker)
            child = {data_child.iloc[i, marker_col],
                     data_child.iloc[i, marker_col+1]}

            # 子のマーカーの集合(要素数2)に対し，母親のマーカーを除く．差集合が空集合の時は，子が母親と完全一致している，もしくは子が1つしか値を持たないということになるため，そのままでよい．
            mother = {data_mother.iloc[i, marker_col],
                      data_mother.iloc[i, marker_col+1]}
            if (child - mother):
                child = child - mother

            # 親候補のループ
            for j in range(data_father.shape[0]):
                marker_col = data_father.columns.get_loc(marker)
                father = {data_father.iloc[j, marker_col],
                          data_father.iloc[j, marker_col+1]}

                # 子(母親のマーカーを除外)と父候補の集合に共通部分があれば父親候補として認める
                if child & father:
                    result[marker].append(1)
                else:
                    result[marker].append(0)

        # 各父親に対して，一致したマーカー数の確認
        summed_values = list(map(sum, zip(*result.values())))
        result['Sum'] = summed_values

        result['Name'] = list(data_father['name'])
        result_df = pd.DataFrame(result)
        result_df.set_index('Name', inplace=True)

        results.append(result_df)
    return results


# エクセルのバイナリデータを返す関数．子の個体番号をエクセルのsheet名とする．
def make_excel_result(df_data, child_names, num_permissible):
    candidates = [[] for _ in range(len(child_names))]
    # BytesIOオブジェクトを使用してメモリ上でExcelファイルを作成
    excel_data = BytesIO()
    with pd.ExcelWriter(excel_data, engine='openpyxl') as writer:
        for i, (df, child_name) in enumerate(zip(df_data, child_names)):
            df.to_excel(writer, sheet_name=child_name, index=True)

            # openpyxlでシートを取得
            worksheet = writer.sheets[child_name]

            # 'Sum'カラムのインデックスを取得 (Excelの列は1から始まるので調整)
            sum_column_index = df.columns.get_loc(
                'Sum') + 2  # pandasのインデックスは0から始まり，インデックスを含むため+2

            # 条件を満たすセルに色を付ける
            for row in range(2, len(df) + 2):  # Excelでは行が1から始まり、ヘッダーがあるため+2
                cell = worksheet.cell(row=row, column=sum_column_index)
                if cell.value >= num_permissible:
                    candidates[i].append(
                        worksheet.cell(row=row, column=1).value)
                    for col in range(1, len(df.columns) + 2):
                        cell_ = worksheet.cell(row=row, column=col)
                        cell_.fill = PatternFill(
                            start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # メモリバッファの位置を先頭に戻す
    excel_data.seek(0)

    return excel_data, candidates


def check_mother(data_child, data_mother):
    markers = [col for col in data_child.columns[1:] if not 'Unnamed' in col]
    results = []

    # 子供の数分ループ
    for i in range(data_child.shape[0]):
        result = {}

        # 各マーカーについて探索
        for marker in markers:
            result[marker] = []
            marker_col = data_child.columns.get_loc(marker)
            child = {data_child.iloc[i, marker_col],
                     data_child.iloc[i, marker_col+1]}

            # 母親候補のループ
            for j in range(data_mother.shape[0]):
                marker_col = data_mother.columns.get_loc(marker)
                mother = {data_mother.iloc[j, marker_col],
                          data_mother.iloc[j, marker_col+1]}

                # 子と母候補の集合に共通部分があれば母親候補として認める
                if child & mother:
                    result[marker].append(1)
                else:
                    result[marker].append(0)

        # 各母親に対して，一致したマーカー数の確認
        summed_values = list(map(sum, zip(*result.values())))
        result['Sum'] = summed_values

        result['Name'] = list(data_mother['name'])
        result_df = pd.DataFrame(result)
        result_df.set_index('Name', inplace=True)

        results.append(result_df)
    return results


# 母親候補の解析結果を，父親の解析に利用可能な形に保存する．
def make_excel_data(df_child, df_mother, candidates):
    child_names = df_child['name'].tolist()

    new_df_child = pd.DataFrame(columns=df_child.columns)
    new_df_mother = pd.DataFrame(columns=df_mother.columns)

    for i, candidate in enumerate(candidates):
        child_row = df_child[df_child['name'] == child_names[i]]
        for j, mother in enumerate(candidate):
            child_row['name'] = child_names[i] + f'_{j:03d}'
            mother_row = df_mother[df_mother['name'] == mother]

            new_df_child = pd.concat(
                [new_df_child, child_row], ignore_index=True)
            new_df_mother = pd.concat(
                [new_df_mother, mother_row], ignore_index=True)

    new_df_child.set_index('name', inplace=True)
    new_df_mother.set_index('name', inplace=True)

    excel_data = BytesIO()
    with pd.ExcelWriter(excel_data, engine='openpyxl') as writer:
        new_df_child.to_excel(writer, sheet_name='child', index=True)
        new_df_mother.to_excel(writer, sheet_name='mother', index=True)
    excel_data.seek(0)

    return excel_data
