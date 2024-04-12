import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO

import time
import method


st.set_page_config(
    page_title="Father Identifying",
    page_icon="🐧",
    layout="centered",
    initial_sidebar_state='auto',
    menu_items={
        'About': "This is an application that identifies individuals who may be the biological fathers of children who may have EPP."
    }
)

st.title('Father Identifying')

st.sidebar.title('パラメータの設定')

# st.sidebar.write('# ')

# st.sidebar.write('## 子の数')
# num_child = st.sidebar.slider('確認する子の数', 1, 30, 1)

st.sidebar.write('# ')

st.sidebar.write('## マーカー数')
num_marker = st.sidebar.slider('使用するマーカー数', 1, 30, 10)

st.sidebar.write('# ')

st.sidebar.write('## 許容度')
num_permissible = st.sidebar.slider(
    '何マーカー以上が一致すれば父親と認めるか', 0, num_marker, num_marker//2)

'# '

st.subheader('テンプレート', divider='gray')
'> 以下のボタンより，テンプレートをダウンロードできます. '

with open('Template_files/template_data_f.xlsx', 'rb') as file:
    st.download_button('ダウンロード', data=file.read(),
                       file_name='template_data.xlsx')

'# \n# '

st.subheader('データのアップロード', divider='gray')

col1, col2 = st.columns([7, 1])
col1.write('> テンプレートのフォーマットに従い, Excelファイルをアップロードしてください.')
col2.write('')
with col2.popover('詳細'):
    st.info('''
            ダウンロードしたExcelファイルのフォーマットに従ってください.\n
            子, 実母親, 父親候補のデータを同一のExcelファイルのsheetに保存してください.\n
            sheet名はそれぞれ"child", "mother", "father" としてください.\n
            注）子のデータと実母親のデータは順番を対応させてください.''')

df_child = df_mother = df_father = None
uploaded_file = st.file_uploader("Excelファイルを選択してください", type=[
                                 "xlsx"], key="file_uploader_1")
if uploaded_file is not None:

    # アップロードされたエクセルファイルを開き，sheet名を取得.
    bytes_data = BytesIO(uploaded_file.read())
    wb = load_workbook(filename=bytes_data, read_only=True)
    sheet_names = wb.sheetnames

    check_list = ['child', 'mother', 'father']

    if all(s in sheet_names for s in check_list):
        df_child = pd.read_excel(uploaded_file, sheet_name='child')
        df_mother = pd.read_excel(uploaded_file, sheet_name='mother')
        df_father = pd.read_excel(uploaded_file, sheet_name='father')

        tab1, tab2, tab3 = st.tabs(
            ['child_data', 'mother_data', 'father_data'])
        with tab1:
            st.write('##### 子のデータ')
            st.write(
                f'> データ数 : {df_child.shape[0]} ,  マーカー数 : {(df_child.shape[1]-1)//2}')
            st.dataframe(df_child.head(), hide_index=True)
        with tab2:
            st.write('##### 実母親のデータ')
            st.write(
                f'> データ数 : {df_mother.shape[0]} ,  マーカー数 : {(df_mother.shape[1]-1)//2}')
            st.dataframe(df_mother.head(), hide_index=True)
        with tab3:
            st.write('##### 父親候補のデータ')
            st.write(
                f'> データ数 : {df_father.shape[0]} ,  マーカー数 : {(df_father.shape[1]-1)//2}')
            st.dataframe(df_father.head(), hide_index=True)
    else:
        st.error('Error : sheet名を確認してください.')


# try:

'# '
st.subheader('親子鑑定の実施', divider='gray')

analyze = st.button('解析する')

if analyze:
    if all(df is not None for df in [df_child, df_mother, df_father]):

        with st.spinner('Analyzing...'):
            time.sleep(1)
            results = method.check_father(df_child, df_mother, df_father)

            child_names = list(df_child['name'])
            excel_data, father_candidates = method.make_excel_result(
                results, child_names, num_permissible)

        st.write('解析が完了しました. \n# ')
        st.subheader('結果', divider='gray')

        with st.expander('See Results', expanded=True):
            for child_name, father_candidate in zip(child_names, father_candidates):
                candidate_str = ", ".join(
                    father_candidate) if father_candidate else 'なし'
                st.markdown(f"""
                    <div style="background-color:#f0f2f6;padding:10px;border-radius:10px;">
                        <h4>{child_name}</h4>
                        <p>父親候補：{candidate_str}</p>
                    </div>
                """, unsafe_allow_html=True)

        '# '

        st.write('**結果は以下のボタンからダウンロード可能です.**')
        st.download_button(
            label='結果をダウンロード',
            data=excel_data,
            file_name='result.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    else:
        st.error('Error : Excelファイルをアップロードしてください.')

# except:
#     st.error('エラーが起こりました. Excelファイルのフォーマットを確認してください. ')
