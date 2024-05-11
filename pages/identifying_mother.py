import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO

import time
import method


st.set_page_config(
    page_title="Mother Identifying",
    page_icon="🐧",
    layout="centered",
    initial_sidebar_state='auto',
    menu_items={
        'About': "This is an application that identifies individuals who may be the biological fathers of children who may have EPP."
    }
)

# ページリンク
st.sidebar.title('ページリンク')
st.sidebar.page_link("epp_checker.py", label="Home", icon="🐧")
st.sidebar.page_link("pages/identifying_father.py",
                     label="Identifying the father", icon="♂️")
st.sidebar.page_link("pages/identifying_mother.py",
                     label="Identifying the mother", icon="♀️", disabled=True)
st.sidebar.divider()

st.sidebar.title('パラメータの設定')

# st.sidebar.write('# ')

# st.sidebar.write('## 子の数')
# num_child = st.sidebar.slider('確認する子の数', 1, 30, 1)

st.sidebar.write('# ')

st.sidebar.write('## マーカー数')
num_marker = st.sidebar.slider('使用するマーカー数', 1, 30, 14)

st.sidebar.write('# ')

st.sidebar.write('## 許容度')
num_permissible = st.sidebar.slider(
    '何マーカー以上が一致すれば母親と認めるか', 0, num_marker, num_marker//2)

st.title('Identifying the Mother')

'# '

st.subheader('テンプレート', divider='gray')
'> 以下のボタンより，テンプレートをダウンロードできます. '

with open('Template_files/template_data_m.xlsx', 'rb') as file:
    st.download_button('Download', data=file.read(),
                       file_name='template_data.xlsx')

'# \n# '

st.subheader('データのアップロード', divider='gray')

col1, col2 = st.columns([7, 1])
col1.write('> テンプレートのフォーマットに従い, Excelファイルをアップロードしてください.')
col2.write('')
with col2.popover('詳細'):
    st.info('''
            ダウンロードしたExcelファイルのフォーマットに従ってください.\n
            子, 母親候補のデータを同一のExcelファイルのsheetに保存してください.\n
            sheet名はそれぞれ"child", "mother"としてください.''')

df_child = df_mother = None
uploaded_file = st.file_uploader("Excelファイルを選択してください", type=[
                                 "xlsx"], key="file_uploader_1")

if uploaded_file is not None:

    # アップロードされたエクセルファイルを開き，sheet名を取得．
    bytes_data = BytesIO(uploaded_file.read())
    wb = load_workbook(filename=bytes_data, read_only=True)
    sheet_names = wb.sheetnames

    check_list = ['child', 'mother']

    if all(s in sheet_names for s in check_list):
        df_child = pd.read_excel(uploaded_file, sheet_name='child')
        df_mother = pd.read_excel(uploaded_file, sheet_name='mother')

        tab1, tab2 = st.tabs(['child_data', 'mother_data'])
        with tab1:
            st.write('##### 子のデータ')
            st.write(
                f'> データ数 : {df_child.shape[0]} ,  マーカー数 : {(df_child.shape[1]-1)//2}')
            st.dataframe(df_child.head(), hide_index=True)
        with tab2:
            st.write('##### 母親候補のデータ')
            st.write(
                f'> データ数 : {df_mother.shape[0]} ,  マーカー数 : {(df_mother.shape[1]-1)//2}')
            st.dataframe(df_mother.head(), hide_index=True)
    else:
        st.error('Error : sheet名を確認してください. ')


# try:

'# '
st.subheader('親子鑑定の実施', divider='gray')

analyze = st.button('Analyze')

if analyze:
    if all(df is not None for df in [df_child, df_mother]):

        with st.spinner('Analyzing...'):
            time.sleep(1)
            results = method.check_mother(df_child, df_mother)

            child_names = list(df_child['name'])
            excel_data_result, mother_candidates = method.make_excel_result(
                results, child_names, num_permissible)

            excel_data = method.make_excel_data(
                df_child, df_mother, mother_candidates)

        st.write('解析が完了しました. \n# ')
        st.subheader('結果', divider='gray')

        with st.expander('See Results', expanded=True):
            for child_name, mother_candidate in zip(child_names, mother_candidates):
                candidate_str = ", ".join(
                    mother_candidate) if mother_candidate else 'なし'
                st.markdown(f"""
                    <div style="background-color:#f0f2f6;padding:10px;border-radius:10px;">
                        <h4>{child_name}</h4>
                        <p>母親候補：{candidate_str}</p>
                    </div>
                """, unsafe_allow_html=True)

        '# '

        st.write('**結果は以下のボタンからダウンロード可能です.**')
        st.download_button(
            label='Download results',
            data=excel_data_result,
            file_name='result.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        st.write('**子，母親のデータは以下のボタンからダウンロード可能です.**')
        st.download_button(
            label='Download new data',
            data=excel_data,
            file_name='result_data.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    else:
        st.error('Error : Excelファイルをアップロードしてください.')

# except:
#     st.error('エラーが起こりました. Excelファイルのフォーマットを確認してください. ')
